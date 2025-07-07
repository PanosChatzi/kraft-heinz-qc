import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
from pathlib import Path
import tempfile

# Built-in mappings configuration
BUILT_IN_MAPPINGS = {
    'APROTEICI': {
        'geographies': [
            'Hypermarkets (7011)',
            'SSS (7013)',
            'Supermarkets (7012)',
            'Total Generalist Online (6100)',
            'Total Italy (inc. Discount) (7406)',
            'Total Italy + Pharma (4380)',
            'Total Italy Hyper+Super+Pharma (4301)',
            'Total Italy Pharma (3930)',
            'Total Italy+Pharma+Online (4397)'
        ],
        'prod_hier_filter': '01-APROTEICO_3'
    },
    'INFANZIA': {
        'geographies': [
            'Discount (58)',
            'Hypermarkets (7011)',
            'SSS (7013)',
            'Supermarkets (7012)',
            'Total Generalist Online (6100)',
            'Total Italy (inc. Discount) (7406)',
            'Total Italy + Pharma (4380)',
            'Total Italy Hyper+Super+Pharma (4301)',
            'Total Italy Pharma (3930)',
            'Total Italy+Pharma+Online (4397)',
            'Traditionals (incl. Microm. <100mq) (7425)'
        ],
        'prod_hier_filter': '02-CATEGORY_1'
    },
    'Sauces': {
        'geographies': [
            'Discount (58)',
            'Hypermarkets (7011)',
            'SSS (7013)',
            'Supermarkets (7012)',
            'Total Generalist Online (6100)',
            'Total Italy (inc. Discount) (7406)',
            'Traditionals (incl. Microm. <100mq) (7425)'
        ],
        'prod_hier_filter': '01-CATEGORY_8'
    },
    'SALSE': {
        'geographies': [
            'Discount (58)',
            'Hypermarkets (7011)',
            'SSS (7013)',
            'Supermarkets (7012)',
            'Total Generalist Online (6100)',
            'Total Italy (inc. Discount) (7406)',
            'Traditionals (incl. Microm. <100mq) (7425)'
        ],
        'prod_hier_filter': '02-CATEGORY_1'
    },
    'GLUTINE': {
        'geographies': [
            'Discount (58)',
            'Hypermarkets (7011)',
            'SSS (7013)',
            'Supermarkets (7012)',
            'Total Generalist Online (6100)',
            'Total Italy (inc. Discount) (7406)',
            'Total Italy + Pharma (4380)',
            'Total Italy Hyper+Super+Pharma (4301)',
            'Total Italy Pharma (3930)',
            'Total Italy+Pharma+Online (4397)',
            'Traditionals (incl. Microm. <100mq) (7425)'
        ],
        'prod_hier_filter': '01-SENZA_GLUTINE_4'
    }
}

class StreamlitExcelComparator:
    def __init__(self, threshold=0.01):
        self.threshold = threshold
        self.mappings = BUILT_IN_MAPPINGS
        self.summary_stats = {
            'compared_sheets': 0,
            'total_mismatches': 0,
            'total_new_dates': 0
        }

    def match_filename_to_mapping(self, filename):
        """Match filename to appropriate mapping configuration (case-insensitive)"""
        filename_base = Path(filename).stem.upper()
        
        # Try case-insensitive match against all patterns
        for original_pattern in self.mappings.keys():
            pattern_upper = original_pattern.upper()
            if pattern_upper in filename_base:
                return original_pattern
        
        return None

    def validate_file_mappings(self, file1_name, file2_name):
        """Validate that both files match the same mapping pattern"""
        pattern1 = self.match_filename_to_mapping(file1_name)
        pattern2 = self.match_filename_to_mapping(file2_name)
        
        # If either file doesn't match any pattern
        if pattern1 is None or pattern2 is None:
            missing_files = []
            if pattern1 is None:
                missing_files.append(f"File 1: {Path(file1_name).name}")
            if pattern2 is None:
                missing_files.append(f"File 2: {Path(file2_name).name}")
            
            error_msg = f"❌ No mapping found for: {', '.join(missing_files)}\n"
            error_msg += "Available mapping patterns:\n"
            for i, pattern in enumerate(self.mappings.keys(), 1):
                error_msg += f"  {i}. {pattern}\n"
            
            return None, None, error_msg
        
        # If files match different patterns
        if pattern1 != pattern2:
            error_msg = f"❌ You have picked different files for comparison:\n"
            error_msg += f"   File 1 pattern: {pattern1}\n"
            error_msg += f"   File 2 pattern: {pattern2}\n"
            error_msg += "Please re-run the script with correct pairs for comparison."
            return None, None, error_msg
        
        return pattern1, pattern1, None

    def filter_extract_by_prodhier(self, df, prod_hier_filter):
        """Filter Extract file by ProdHier column"""
        if 'ProdHier' not in df.columns:
            return df

        filtered_df = df[df['ProdHier'] == prod_hier_filter].copy()

        if filtered_df.empty:
            st.warning("Warning: select the correct UNIFY and EXTRACT files.")

        return filtered_df

    def parse_dates(self, df, col_name):
        """Convert date columns to datetime"""
        if col_name not in df.columns:
            return df

        try:
            if df[col_name].dtype == 'object' and 'Week ending' in str(df[col_name].iloc[0]):
                df['Date'] = pd.to_datetime(
                    df[col_name].str.extract(r'(\d{2}-\d{2}-\d{4})')[0],
                    dayfirst=True
                )
            else:
                df['Date'] = pd.to_datetime(df[col_name])
            return df
        except Exception as e:
            st.error(f"Error parsing dates from column '{col_name}': {e}")
            return df

    def compare_values(self, val1, val2):
        """Compare two values and return difference if above threshold"""
        if pd.isna(val1) and pd.isna(val2):
            return None, 0
        if pd.isna(val1) or pd.isna(val2):
            return True, abs(float(val1 or 0) - float(val2 or 0))

        try:
            v1, v2 = float(val1), float(val2)
            diff = abs(v1 - v2)
            return diff > self.threshold, diff
        except (ValueError, TypeError):
            return val1 != val2, 0

    def get_column_mapping(self, mapping_pattern, df1_cols, df2_cols):
        """Create column mapping between two dataframes based on pattern (case-insensitive)"""
        if mapping_pattern not in self.mappings:
            return {}

        target_columns = self.mappings[mapping_pattern]['geographies']
        column_map = {}

        for target_col in target_columns:
            col1_match = None
            col2_match = None

            # Find matching columns (case-insensitive)
            for col in df1_cols:
                if (target_col.lower() == str(col).lower() or
                    target_col.lower() in str(col).lower() or
                    str(col).lower() in target_col.lower()):
                    col1_match = col
                    break

            for col in df2_cols:
                if (target_col.lower() == str(col).lower() or
                    target_col.lower() in str(col).lower() or
                    str(col).lower() in target_col.lower()):
                    col2_match = col
                    break

            if col1_match and col2_match:
                column_map[col1_match] = col2_match
            else:
                st.warning(f"⚠️  No match found for target column: '{target_col}'")

        return column_map

    def find_date_column(self, df):
        """Find date column in dataframe"""
        date_columns_to_try = ['Time', 'Date', 'PER_DESCRIPTION', 'Week ending', 'Period']

        for col in date_columns_to_try:
            if col in df.columns:
                return col
        return None

    def generate_report(self, mismatch_rows, common_dates, mapping_pattern,
                        total_comparisons, significant_differences, file1_name, file2_name,
                        prod_hier_filter, new_dates):
        """Generate Excel report with results"""
        first_date = common_dates.min().strftime('%Y-%m-%d')
        last_date = common_dates.max().strftime('%Y-%m-%d')

        # Create a temporary file for the report
        current_date = pd.Timestamp.now().strftime('%d-%m-%Y')
        
        # Create a BytesIO buffer for the Excel file
        from io import BytesIO
        buffer = BytesIO()

        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Summary sheet
            summary_data = {
                "Metric": [
                    "First Date", "Last Date", "Mapping Pattern", "ProdHier Filter",
                    "File 1", "File 2", "Total Comparisons", "Significant Differences",
                    "New Dates Count", "Threshold Used", "Success Rate %"
                ],
                "Value": [
                    first_date, last_date, mapping_pattern, prod_hier_filter,
                    Path(file1_name).name, Path(file2_name).name, total_comparisons,
                    significant_differences, len(new_dates), self.threshold,
                    round((total_comparisons - significant_differences) / total_comparisons * 100, 2) if total_comparisons > 0 else 100
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, index=False, sheet_name="Summary")

            # Significant differences sheet
            if mismatch_rows:
                differences_df = pd.DataFrame(mismatch_rows)
                differences_df.to_excel(writer, index=False, sheet_name="Significant_Differences")
                self.highlight_differences_in_writer(writer, "Significant_Differences")
            else:
                pd.DataFrame([{"Message": "No mismatches found."}]).to_excel(writer, index=False, sheet_name="Mismatches")

            # New dates sheet
            if len(new_dates) > 0:
                new_dates_df = pd.DataFrame({
                    'New_Dates': [date.strftime('%Y-%m-%d') for date in new_dates]
                })
                new_dates_df.to_excel(writer, index=False, sheet_name="New_Dates")

        buffer.seek(0)
        return buffer, f"qc_report_{mapping_pattern}_{current_date}.xlsx"

    def highlight_differences_in_writer(self, writer, sheet_name):
        """Highlight significant differences in the Excel writer"""
        try:
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            # Highlight rows with differences above threshold
            for row in range(2, worksheet.max_row + 1):  # Skip header
                diff_cell = worksheet.cell(row=row, column=6)  # Difference column
                if diff_cell.value and float(diff_cell.value) > self.threshold:
                    for col in range(1, worksheet.max_column + 1):
                        worksheet.cell(row=row, column=col).fill = fill
        except Exception as e:
            st.error(f"Error applying highlighting: {e}")

    def compare_files(self, unify_file, extract_file):
        """Main comparison function for Streamlit"""
        try:
            # Validate that both files match the same mapping pattern
            pattern1, pattern2, error_msg = self.validate_file_mappings(unify_file.name, extract_file.name)
            
            if pattern1 is None or pattern2 is None:
                return None, error_msg

            mapping_pattern = pattern1
            prod_hier_filter = self.mappings[mapping_pattern]['prod_hier_filter']

            # Load dataframes
            df1 = pd.read_excel(unify_file, skiprows=7)            # Unify
            df2 = pd.read_excel(extract_file, sheet_name='Sheet2')    # Extract

            # Filter Extract file by ProdHier
            df2_filtered = self.filter_extract_by_prodhier(df2, prod_hier_filter)

            if df2_filtered.empty:
                return None, f"No data remains after filtering by ProdHier = '{prod_hier_filter}'"

            # Parse dates
            df1_date_col = self.find_date_column(df1)
            df2_date_col = self.find_date_column(df2_filtered)

            if not df1_date_col or not df2_date_col:
                return None, "Could not find date columns in one or both files"

            df1 = self.parse_dates(df1, df1_date_col)
            df2_filtered = self.parse_dates(df2_filtered, df2_date_col)

            if 'Date' not in df1.columns or 'Date' not in df2_filtered.columns:
                return None, "Could not parse dates from files"

            # Set date as index
            df1.set_index('Date', inplace=True)
            df2_filtered.set_index('Date', inplace=True)

            # Find common dates and new dates
            common_dates = df1.index.intersection(df2_filtered.index)
            new_dates = df2_filtered.index.difference(df1.index)

            if len(common_dates) == 0:
                return None, "No common dates found between files"

            df1_aligned = df1.loc[common_dates]
            df2_aligned = df2_filtered.loc[common_dates]

            # Get column mapping
            column_map = self.get_column_mapping(mapping_pattern, df1_aligned.columns, df2_aligned.columns)

            if not column_map:
                return None, "No column mappings found. Please check your mapping configuration."

            # Compare mapped columns
            mismatch_rows = []
            total_comparisons = 0
            significant_differences = 0

            for col1, col2 in column_map.items():
                for date in common_dates:
                    val1 = df1_aligned.at[date, col1]
                    val2 = df2_aligned.at[date, col2]
                    total_comparisons += 1

                    is_different, difference = self.compare_values(val1, val2)

                    if is_different and difference > self.threshold:
                        significant_differences += 1
                        mismatch_rows.append({
                            "Date": date.strftime('%Y-%m-%d'),
                            "Column File 1": col1,
                            "Value File 1": val1,
                            "Column File 2": col2,
                            "Value File 2": val2,
                            "Difference": difference,
                            "Above_Threshold": True
                        })

            # Update summary stats
            self.summary_stats['compared_sheets'] = 1
            self.summary_stats['total_mismatches'] = significant_differences
            self.summary_stats['total_new_dates'] = len(new_dates)

            # Generate report
            report_buffer, report_filename = self.generate_report(
                mismatch_rows, common_dates, mapping_pattern,
                total_comparisons, significant_differences, 
                unify_file.name, extract_file.name,
                prod_hier_filter, new_dates
            )

            # Create summary for display
            first_date = common_dates.min().strftime('%Y-%m-%d')
            last_date = common_dates.max().strftime('%Y-%m-%d')
            
            summary = {
                'first_date': first_date,
                'last_date': last_date,
                'compared_sheets': self.summary_stats['compared_sheets'],
                'total_mismatches': self.summary_stats['total_mismatches'],
                'total_new_dates': self.summary_stats['total_new_dates'],
                'mapping_pattern': mapping_pattern,
                'success_rate': round((total_comparisons - significant_differences) / total_comparisons * 100, 2) if total_comparisons > 0 else 100
            }

            return {
                'summary': summary,
                'report_buffer': report_buffer,
                'report_filename': report_filename
            }, None

        except Exception as e:
            return None, f"Error during comparison: {str(e)}"

def main():
    st.set_page_config(page_title="Kraft Heinz QC", page_icon="📊", layout="wide")
    
    st.title("📊 Kraft Heinz QC")
    st.markdown("Compare UNIFY and EXTRACT excel files.")
    
    # Initialize session state
    if 'comparison_result' not in st.session_state:
        st.session_state.comparison_result = None
    if 'error_message' not in st.session_state:
        st.session_state.error_message = None
    
    # File upload section
    st.header("📁 Upload files")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("UNIFY file")
        unify_file = st.file_uploader(
            "Choose UNIFY excel file (e.g., 1. Check_Model_Custom_APROTEICI (33).xlsx)",
            type=['xlsx', 'xls'],
            key="unify_file"
        )
        if unify_file:
            st.success(f"✅ {unify_file.name}")
    
    with col2:
        st.subheader("EXTRACT file")
        extract_file = st.file_uploader(
            "Choose EXTRACT excel file (e.g., IT_2393_APROTEICI_D156.xlsx)",
            type=['xlsx', 'xls'],
            key="extract_file"
        )
        if extract_file:
            st.success(f"✅ {extract_file.name}")
    
    # Comparison button
    st.header("🔍 Comparison")
    
    if st.button("Compare files", type="primary", disabled=not (unify_file and extract_file)):
        if unify_file and extract_file:
            with st.spinner("Comparing files..."):
                comparator = StreamlitExcelComparator()
                result, error = comparator.compare_files(unify_file, extract_file)
                
                if error:
                    st.session_state.error_message = error
                    st.session_state.comparison_result = None
                else:
                    st.session_state.comparison_result = result
                    st.session_state.error_message = None
    
    # Display results
    if st.session_state.error_message:
        st.error(st.session_state.error_message)
    
    if st.session_state.comparison_result:
        st.header("📈 Summary Statistics")
        
        summary = st.session_state.comparison_result['summary']
        
        # Display summary in a nice format
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.metric("First Date", summary['first_date'])
            st.metric("Last Date", summary['last_date'])
        
        with col2:
            st.metric("Compared Sheets", summary['compared_sheets'])
            st.metric("Total Mismatches", summary['total_mismatches'])
        
        with col3:
            st.metric("Total New Dates", summary['total_new_dates'])
            st.metric("Success Rate", f"{summary['success_rate']}%")
                
        # Download button
        st.header("📄 Download Report")
        
        st.download_button(
            label="Download Excel Report",
            data=st.session_state.comparison_result['report_buffer'],
            file_name=st.session_state.comparison_result['report_filename'],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        st.success("Report generated successfully! Click the button above to download.")

if __name__ == "__main__":
    main()